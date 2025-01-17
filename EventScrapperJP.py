import requests
import re
from bs4 import BeautifulSoup
import openpyxl
import pykakasi
import pandas as pd
import time
import os
from openpyxl.styles import Font, Color, PatternFill, Alignment, Fill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import helper
EXCEL_FILE = "EventsJP2025.xlsx"

def PiaInnerScrapper(url):
    Inner_doc_Pia = helper.doc_from_url(url)
    for div_Inner_Pia in Inner_doc_Pia.find_all("div", class_="textDefinitionList-2024__item"):
        dt_tag_Pia = div_Inner_Pia.find("dt", class_="textDefinitionList-2024__title")
        if dt_tag_Pia and "会場" in dt_tag_Pia.get_text(strip=True):
            dd_tag_Pia = div_Inner_Pia.find("dd", class_="textDefinitionList-2024__desc")
            if dd_tag_Pia:
                return dd_tag_Pia.get_text(strip=True)
                #if PPia:
                #    return convert_to_romaji(PPia)
                #else:
                #    return None   

def PiaScrapper(doc_Pia):
    i=0
    Piaconcerts = []
    for div_Pia in doc_Pia.find_all("div"):
        a_tag_Pia = div_Pia.find("a")
        
        if a_tag_Pia:
            figcaption_Pia = a_tag_Pia.find("figcaption")
            
            if figcaption_Pia:
                
                name_Pia = figcaption_Pia.find("h2")
                namePia = name_Pia.get_text(strip=True) if name_Pia else None
                
                if namePia:
                    romajiPia = helper.convert_to_romaji(namePia)
                else:
                    romajiPia = None
                    
                date_Pia= figcaption_Pia.find("span")
                datePia = date_Pia.get_text(strip=True) if date_Pia else None
                
                linkPia = a_tag_Pia.get("href") if a_tag_Pia else None
                               
                placePia = PiaInnerScrapper(linkPia)
                                 
                if namePia and datePia and linkPia:
                    if not any(linkPia in Piaconcert["Link"] for Piaconcert in Piaconcerts):
                        Piaconcerts.append({"Name": namePia, "Romaji": romajiPia, "Place": placePia, "Date": datePia, "Link": linkPia})
                        i+=1
                        if i>1:
                            break #tester
                        print(i)                       
    print(f"Finished scraping current site. Proceeding to the next one.")
    return Piaconcerts

def eplusScrapper(doc_eplus, month): #fix the dates edge case
    i=0
    Eplusconcerts = []
    li_eplus = doc_eplus.find("li", class_="block-paginator__item block-paginator__item--last")
    if li_eplus:
        max_pages = int(li_eplus.get_text(strip=True))
        print(f"Max pages: {max_pages}")
    
        for page in range(1, max_pages+1):
            url = f"https://eplus.jp/sf/event/month-0{month}/p{page}"
            print(f"Scraping page: {url}")
            doc_eplus = helper.doc_from_url(url)
            
            ticket_div = doc_eplus.find("div", class_="block-ticket-list__content output")
            tickets = ticket_div.find_all("a")

            for ticket in tickets:
                nameEplus = (ticket.find("h3", class_="ticket-item__title")).get_text(strip=True)
                
                if nameEplus:
                    romajiEplus = helper.convert_to_romaji(nameEplus)
                else:
                    romajiEplus = None
                
                date_year = ticket.find("span", class_="ticket-item__yyyy")
                date_mmdd = ticket.find("span", class_="ticket-item__mmdd")
                dateEplus = (f"{date_year.get_text(strip=True)}{date_mmdd.get_text(strip=True)}" 
                             if date_year and date_mmdd else None)
                div_eplus_venue = ticket.find("div", class_="ticket-item__venue")
                if div_eplus_venue:
                    place_eplus = div_eplus_venue.find("p")
                    placeEplus = place_eplus.get_text(strip=True) 
                    linkEplus = "https://eplus.jp" + ticket.get("href") if ticket else None
                    if nameEplus and dateEplus and linkEplus:
                        if not any(linkEplus in Eplusconcert["Link"] for Eplusconcert in Eplusconcerts):
                            Eplusconcerts.append({"Name": nameEplus, "Romaji": romajiEplus, "Place": placeEplus, "Date": dateEplus, "Link": linkEplus})
                            i+=1
                            #if i>1:
                            #    break #tester
                            print(i)
    print(f"Finished scraping current site. Proceeding to the next one.")
    return Eplusconcerts

def remove_duplicates_in_excel_pia():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Events_t.pia.jp"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    unique_rows = [headers]

    seen = set()
    for row in rows[1:]:
        link = row[4] #Change if columns moved
        if link not in seen:
            unique_rows.append(row)
            seen.add(link)

    sheet.delete_rows(1, sheet.max_row)
    for unique_row in unique_rows:
        sheet.append(unique_row)

    helper.save_workbook(workbook)
    
def remove_duplicates_in_excel_eplus():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook["Events_eplus.jp"]
    
    seen = {}
    
    row_number = 2
    while row_number <= sheet.max_row:
        name = sheet.cell(row=row_number, column=1).value
        ending_date = sheet.cell(row=row_number, column=5).value
        
        if name in seen:
            original_row = seen[name]
            original_ending_date = sheet.cell(row=original_row, column=5).value
            if ending_date > original_ending_date:
                sheet.cell(row=original_row, column=5).value = ending_date
            sheet.delete_rows(row_number)
        else:
            seen[name] = row_number
            row_number += 1
    
    helper.save_workbook(workbook)
    
def style_sort_excel(sheet_name, sorting_column):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    df = pd.DataFrame(sheet.values)
    headers = df.iloc[0]
    df = df[1:]
    df.columns = headers
    df = df.sort_values(by=sorting_column, ascending=True)
    sheet.delete_rows(1, sheet.max_row)
    for row in [df.columns.tolist()] + df.values.tolist():
       sheet.append(row)
    
    title_row_style = Font(size=14, color="FFFFFF", bold=True)
    for i in range (0,column_count):
        sheet.cell(row=1, column=i+1).font = title_row_style
    dim_holder = DimensionHolder(worksheet=sheet)
    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=35)
    sheet.column_dimensions = dim_holder
    
    for z in range (0, column_count):
        sheet.cell(row=1, column = z + 1).fill = PatternFill(start_color="38AA49", end_color="38AA49", fill_type="solid")
    for x in range(2, row_count):
        for z in range (0, column_count): 
            c = sheet.cell(row=x, column=z + 1)
            if x % 2 != 0:
                c.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")
    if row_count % 2 != 0:
        for z in range (0, column_count):
            l = sheet.cell(row=row_count, column=z + 1)
            l.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")

doc_PiaM = helper.doc_from_url("https://t.pia.jp/music/")
doc_PiaA = helper.doc_from_url("https://t.pia.jp/anime/")
doc_PiaE = helper.doc_from_url("https://t.pia.jp/event/")

Piaconcerts = PiaScrapper(doc_PiaM) + PiaScrapper(doc_PiaA) + PiaScrapper(doc_PiaE)

sheet_name = "Events_t.pia.jp"
header = ["Name", "Romaji", "Place", "Date", "Link"] #If new column added, change.

workbook, sheet = helper.OpenSheet(sheet_name, header)    
    
for Piaconcert in Piaconcerts:
    sheet.append([Piaconcert["Name"], Piaconcert["Romaji"], Piaconcert["Place"], Piaconcert["Date"], Piaconcert["Link"]])
    
helper.save_workbook(workbook)

remove_duplicates_in_excel_pia()

style_sort_excel(sheet_name, "Date")

print(f"Done! Scraped t.pia.jp. Data saved to {EXCEL_FILE}.")

## Here we start scrapping eplus.jp ##

try:
    doc_eplus_april = helper.doc_from_url("https://eplus.jp/sf/event/month-04")
except:
    print("Error with eplus.jp. It's probably asleep. Trying again later.")
    exit()
try:
    doc_eplus_may = helper.doc_from_url("https://eplus.jp/sf/event/month-05")
except:
    print("Error with eplus.jp. It's probably asleep. Trying again later.")
    exit()

Eplusconcerts = eplusScrapper(doc_eplus_april, 4) + eplusScrapper(doc_eplus_may, 5)

sheet_name = "Events_eplus.jp"
header = ["Name", "Romaji", "Place", "Beginning Date", "Ending Date", "Link"] #If new column added, change.

workbook, sheet = helper.OpenSheet(sheet_name, header)

for Eplusconcert in Eplusconcerts:
    sheet.append([Eplusconcert["Name"], Eplusconcert["Romaji"], Eplusconcert["Place"], Eplusconcert["Date"], Eplusconcert["Date"], Eplusconcert["Link"]])
helper.save_workbook(workbook)

remove_duplicates_in_excel_eplus()

style_sort_excel(sheet_name, "Beginning Date") #maybe a problem with beginning date?

print(f"Done! Scraped eplus.jp. Data saved to {EXCEL_FILE}.")